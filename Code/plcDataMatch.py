import difflib
import pandas as pd
from plcDataMerge import  plcDataMerge
from plcDataProcess import plcDataProcessing
from plcDataTokenization import fenci
from fenciPad import fenciPadding
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

path = os.getcwd()

def string_similar(s1, s2):
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()

def plcAdd(rawData, testData, testDataPred):

    plcDataProcessing(testData)
    fenci(testData)
    fenciPadding(testData)

    plcDataMerge(rawData)
    plcDataProcessing(rawData)
    fenci(rawData)
    fenciPadding(rawData)

    TestData = pd.read_excel(path + '//Files//' + testData + '.xlsx')
    RawDataMerge = pd.read_excel(path + '//Files//' + rawData + 'Merge.xlsx')
    RawDataProcess = pd.read_excel(path + '//Files//' + rawData + 'Process.xlsx')
    RawDataProcess_label = RawDataProcess['Category']
    
    with open(path + '//Files//' + testData + 'FenciPad2.txt', "r") as f:
        #测试数据的fenciPad2表
        TestDataFenciPad2 = f.readlines()
    
    PredIdx = []
    for w in range(len(TestData['Name'])):
    #经典匹配
        if TestData.at[w, 'Name'] in RawDataMerge['Name'].values:
            TestData.at[w, 'Mechanism'] = RawDataMerge[RawDataMerge['Name'] == TestData.at[w, 'Name']].iloc[0,1]
            TestData.at[w, 'Components'] = RawDataMerge[RawDataMerge['Name'] == TestData.at[w, 'Name']].iloc[0,2]
            TestData.at[w, 'InnerMonitoring'] = RawDataMerge[RawDataMerge['Name'] == TestData.at[w, 'Name']].iloc[0,3]
            TestData.at[w, 'Similarity'] = 1
        else:
            #分词匹配
            TestDataFenciPad2[w] = TestDataFenciPad2[w].strip('\n').split() 
            sv = []
            for v in range(len(RawDataProcess_label)):
                sv.append(string_similar(''.join(TestDataFenciPad2[w]), ''.join(RawDataProcess_label[v])))
            idx = sv.index(max(sv))
            TestData.at[w, 'Mechanism'] = RawDataMerge.at[idx, 'Mechanism']
            TestData.at[w, 'Components'] = RawDataMerge.at[idx, 'Components']
            TestData.at[w, 'InnerMonitoring'] = RawDataMerge.at[idx, 'InnerMonitoring']
            TestData.at[w, 'Similarity'] = max(sv)
            PredIdx.append(w+2)

    TestData =TestData.sort_values(by='Similarity')
    TestData.to_excel(path + '//Files//' + testDataPred + '.xlsx', index=False)

    #分词匹配颜色显示
    wb = load_workbook(path + '//Files//' + testDataPred + '.xlsx')
    wb_name = wb.sheetnames
    sheet = wb[wb_name[0]]
    fille = PatternFill('solid', fgColor='00FF0000')
    for i in range(1, len(PredIdx)+1):
        for j in range(1,sheet.max_column+1):
                sheet.cell(row=i+1,column=j).fill=fille

    wb.save(path + '//Files//' + testDataPred + '.xlsx')

    os.remove(path + '//Files//'+rawData+'Merge.xlsx')
    os.remove(path + '//Files//'+rawData+'Process.xlsx')
    os.remove(path + '//Files//'+testData+'Process.xlsx')

    os.remove(path + '//Files//'+rawData+'Process.txt')
    os.remove(path + '//Files//'+rawData+'Fenci.txt')
    os.remove(path + '//Files//'+rawData+'FenciPad.txt')
    os.remove(path + '//Files//'+rawData+'FenciPad2.txt')
    os.remove(path + '//Files//'+testData+'Process.txt')
    os.remove(path + '//Files//'+testData+'Fenci.txt')
    os.remove(path + '//Files//'+testData+'FenciPad.txt')
    os.remove(path + '//Files//'+testData+'FenciPad2.txt')
